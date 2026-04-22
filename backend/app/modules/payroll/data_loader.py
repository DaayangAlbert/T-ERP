from __future__ import annotations

import csv
import json
import logging
import zipfile
from pathlib import Path
from typing import Any, Iterable, Mapping
from xml.etree import ElementTree as ET

from .models import CompanyInfo, EmployeePayrollData, PayrollConfig
from decimal import Decimal

from .utils import normalize_key, normalize_rate, parse_date, parse_decimal, parse_optional_decimal


LOGGER = logging.getLogger(__name__)
SPREADSHEET_NS = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
REL_NS = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}
DOC_RELATIONSHIP = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"


def load_company_info(source: str | Path | Mapping[str, Any]) -> CompanyInfo:
    data = _load_json_like(source)
    company = CompanyInfo.from_dict(data)
    if not company.name:
        raise ValueError("Le fichier company.json doit contenir le nom de l'entreprise.")
    return company


def load_payroll_config(source: str | Path | Mapping[str, Any]) -> PayrollConfig:
    data = _load_json_like(source)
    config = PayrollConfig.from_dict(data, parse_decimal=parse_decimal, normalize_rate=normalize_rate)
    if config.cnps_ceiling is None:
        raise ValueError("Le fichier payroll_config.json doit definir 'cnps_ceiling'.")
    return config


def load_employee_data(
    source: str | Path | Iterable[Mapping[str, Any]],
    logger: logging.Logger | None = None,
) -> list[EmployeePayrollData]:
    active_logger = logger or LOGGER
    rows = _load_rows(source)
    employees: list[EmployeePayrollData] = []

    for index, row in enumerate(rows, start=2):
        employee = _build_employee_record(row, source_row_number=index, logger=active_logger)
        if employee is not None:
            employees.append(employee)

    return employees


def _load_json_like(source: str | Path | Mapping[str, Any]) -> Mapping[str, Any]:
    if isinstance(source, Mapping):
        return source

    path = Path(source)
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def _load_rows(source: str | Path | Iterable[Mapping[str, Any]]) -> list[Mapping[str, Any]]:
    if not isinstance(source, (str, Path)):
        return list(source)

    path = Path(source)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _load_rows_from_csv(path)
    if suffix == ".xlsx":
        return _load_rows_from_xlsx(path)
    if suffix == ".json":
        with path.open("r", encoding="utf-8") as handle:
            content = json.load(handle)
        if isinstance(content, list):
            return [row for row in content if isinstance(row, Mapping)]
        raise ValueError("Le fichier JSON employe doit contenir une liste d'objets.")
    raise ValueError(f"Source non prise en charge: {path}")


def _load_rows_from_csv(path: Path) -> list[Mapping[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def _load_rows_from_xlsx(path: Path) -> list[Mapping[str, Any]]:
    try:
        from openpyxl import load_workbook  # type: ignore

        workbook = load_workbook(path, data_only=True, read_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(value).strip() if value is not None else "" for value in rows[0]]
        return [
            {headers[index]: value for index, value in enumerate(row) if index < len(headers)}
            for row in rows[1:]
            if any(value not in (None, "") for value in row)
        ]
    except ImportError:
        return _load_rows_from_xlsx_stdlib(path)


def _load_rows_from_xlsx_stdlib(path: Path) -> list[Mapping[str, Any]]:
    with zipfile.ZipFile(path) as archive:
        shared_strings = _parse_shared_strings(archive)
        sheet_path = _resolve_first_sheet_path(archive)
        sheet_root = ET.fromstring(archive.read(sheet_path))
        raw_rows: list[list[Any]] = []

        for row in sheet_root.findall(".//m:sheetData/m:row", SPREADSHEET_NS):
            cells: dict[int, Any] = {}
            for cell in row.findall("m:c", SPREADSHEET_NS):
                reference = cell.attrib.get("r", "")
                index = _column_index_from_reference(reference)
                cells[index] = _extract_cell_value(cell, shared_strings)
            if cells:
                max_index = max(cells)
                raw_rows.append([cells.get(index, "") for index in range(max_index + 1)])

    if not raw_rows:
        return []

    headers = [str(value).strip() if value is not None else "" for value in raw_rows[0]]
    return [
        {headers[index]: value for index, value in enumerate(row) if index < len(headers)}
        for row in raw_rows[1:]
        if any(value not in (None, "") for value in row)
    ]


def _parse_shared_strings(archive: zipfile.ZipFile) -> list[str]:
    try:
        root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    except KeyError:
        return []

    strings: list[str] = []
    for string_item in root.findall("m:si", SPREADSHEET_NS):
        parts = [node.text or "" for node in string_item.findall(".//m:t", SPREADSHEET_NS)]
        strings.append("".join(parts))
    return strings


def _resolve_first_sheet_path(archive: zipfile.ZipFile) -> str:
    workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
    sheet = workbook_root.find(".//m:sheets/m:sheet", SPREADSHEET_NS)
    if sheet is None:
        raise ValueError("Le fichier Excel ne contient aucune feuille.")

    rel_id = sheet.attrib.get(f"{{{DOC_RELATIONSHIP}}}id")
    rel_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    for relation in rel_root.findall("r:Relationship", REL_NS):
        if relation.attrib.get("Id") == rel_id:
            target = relation.attrib["Target"].lstrip("/")
            if target.startswith("xl/"):
                return target
            return f"xl/{target}"
    raise ValueError("Impossible de localiser la feuille Excel principale.")


def _column_index_from_reference(reference: str) -> int:
    letters = "".join(character for character in reference if character.isalpha())
    if not letters:
        return 0
    index = 0
    for character in letters:
        index = index * 26 + (ord(character.upper()) - 64)
    return max(index - 1, 0)


def _extract_cell_value(cell: ET.Element, shared_strings: list[str]) -> Any:
    cell_type = cell.attrib.get("t")
    if cell_type == "inlineStr":
        return "".join(node.text or "" for node in cell.findall(".//m:t", SPREADSHEET_NS))

    value = cell.findtext("m:v", default="", namespaces=SPREADSHEET_NS)
    if cell_type == "s" and value:
        return shared_strings[int(value)]
    if cell_type == "b":
        return value == "1"
    return value


def _build_employee_record(
    row: Mapping[str, Any],
    source_row_number: int,
    logger: logging.Logger,
) -> EmployeePayrollData | None:
    normalized = {normalize_key(key): value for key, value in row.items() if key not in (None, "")}

    employee_id = _as_text(normalized.get("employee_id")) or _as_text(normalized.get("matricule"))
    nom = _as_text(normalized.get("nom"))
    prenom = _as_text(normalized.get("prenom"))
    matricule = _as_text(normalized.get("matricule"))

    if not employee_id or not nom or not prenom or not matricule:
        logger.warning(
            "Ligne %s ignoree: colonnes minimales manquantes (employee_id, nom, prenom, matricule).",
            source_row_number,
        )
        return None

    try:
        anciennete = int(parse_decimal(normalized.get("anciennete_mois"), default=0) or 0)
    except (TypeError, ValueError):
        anciennete = 0

    return EmployeePayrollData(
        employee_id=employee_id,
        nom=nom,
        prenom=prenom,
        matricule=matricule,
        categorie=_as_text(normalized.get("categorie")),
        echelon=_as_text(normalized.get("echelon")),
        anciennete_mois=anciennete,
        cnps_number=_as_text(normalized.get("cnps_number")),
        convention_collective=_as_text(normalized.get("convention_collective")),
        emploi=_as_text(normalized.get("emploi")),
        departement=_as_text(normalized.get("departement")),
        date_embauche=parse_date(normalized.get("date_embauche")),
        horaire=_as_text(normalized.get("horaire")),
        situation_famille=_as_text(normalized.get("situation_famille")),
        numero_compte=_as_text(normalized.get("numero_compte")),
        domiciliation=_as_text(normalized.get("domiciliation")),
        jours_payes=parse_decimal(normalized.get("jours_payes"), default=Decimal("0")) or Decimal("0"),
        salaire_base_mensuel=parse_decimal(normalized.get("salaire_base_mensuel"), default=Decimal("0"))
        or Decimal("0"),
        indemn_transport=parse_decimal(normalized.get("indemn_transport"), default=Decimal("0")) or Decimal("0"),
        autres_gains=parse_decimal(normalized.get("autres_gains"), default=Decimal("0")) or Decimal("0"),
        mode_paiement=_as_text(normalized.get("mode_paiement")),
        date_debut_periode=parse_date(normalized.get("date_debut_periode")),
        date_fin_periode=parse_date(normalized.get("date_fin_periode")),
        date_paiement=parse_date(normalized.get("date_paiement")),
        brut_imposable=parse_optional_decimal(normalized.get("brut_imposable")),
        irpp=parse_optional_decimal(normalized.get("irpp")),
        cac=parse_optional_decimal(normalized.get("cac")),
        tc=parse_optional_decimal(normalized.get("tc")),
        rav=parse_optional_decimal(normalized.get("rav")),
        cfs=parse_optional_decimal(normalized.get("cfs")),
        observation=_as_text(normalized.get("observation")),
        source_values=normalized,
        source_row_number=source_row_number,
    )


def _as_text(value: Any) -> str:
    return str(value or "").strip()
